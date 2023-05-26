import openpyxl
import json
from django.conf import settings
from django.template.defaultfilters import slugify
from .models import Provincia, Comarca, Municipi
import os
from django.apps import apps


def run():

    carregats, errors = [], []
    
    # Per cada model, tindrem un fitxer a carregar amb el mateix nom. I l'ordre és important.
    fitxers = ['Provincia','Comarca','Municipi']    

    # Per cada fitxer 
    for fitxer in fitxers:

        carregats_aux, errors_aux = load_excel(fitxer)
        carregats += carregats_aux
        errors += errors_aux

    # Mostrem resultats    
    #mostra_resultats()
    # Tornem els fitxers carregats
    return carregats, errors    
    

def load_excel(fitxer):

    carregats, errors = [], []
    
    file_name = f"{fitxer}.xlsx"

    excel_file = os.path.join(
        settings.BASE_DIR, "geografia/ExcelsUpload/", file_name)
    existeix = os.path.exists(excel_file)

    if settings.DEBUG:
        trobat = "(fitxer trobat)" if existeix else "**No Trobat**"
        print(f"- file_name: {file_name} {trobat}")

    # Si no existeix saltem al següent
    if not existeix:
        return carregats, errors

    wb = openpyxl.load_workbook(filename=excel_file, data_only=True)

    # Carreguem el fitxer al seu corresonent model
    carregats_aux, errors_aux = carrega_fitxer(wb, file_name, fitxer)
    carregats += carregats_aux
    errors += errors_aux
    
    return carregats, errors

def carrega_fitxer(wb, file_name, fitxer):    
   
    carregats, errors = [], []    
   
    # Cerco el model a carregar, és el nom de l'excel a carregar sense l'extensió   
    Model = apps.get_model(app_label='geografia', model_name=fitxer)

    # Esborro dades anteriors d'aquest model
    Model.objects.all().delete()
    
    # Carrego la pestanya. Cada excel té una sola pestanya
    ws = wb.active
    # Nombre de columnes
    n = ws.max_column

    # Creem la capcelera    
    # Recuperem els valors de la capcelera que coincideixen amb els noms del items del model
    capcelera = [cell.value for cell in ws[1][:n]]
      
    # Carreguem l'excel
    # Per cada linia de l'excel 
    for row in list(ws.rows)[1:]:
        if row[0].value is None:
            break
        model = Model()
        cells = [cell for cell in row[:n]]
        # Creem un diccionari amb els valors de les cel.les duplades amb el nom de la cada columna
        data = dict(zip(capcelera, cells))        

        for columna in capcelera:
            value = data[columna].value
            propietat = columna
            
            # El cas del Municipi és diferent. Ja que el codi de comarca i el codi de provincia estan 
            # referenciats als models Comarca i Provincia. Hem de recuperar la instància del model.   
            if fitxer == 'Municipi':
                # busco el codi de comarca al model de comarca
                if columna == 'codi_comarca':

                    propietat = 'codi_comarca_id'
                    valorcomarca = data[columna].value
                    comarcamodel, _ = Comarca.objects.get_or_create(
                    codi=valorcomarca)                    
                    value = comarcamodel.codi
                    
                # busco el codi de província al model de provincia
                elif columna == 'codi_provincia':

                    propietat = 'codi_provincia_id'
                    valorprovincia = data[columna].value
                    provinciamodel, _ = Provincia.objects.get_or_create(
                    codi=valorprovincia)                                        
                    value = provinciamodel.codi

            #if settings.DEBUG:
            #    print( f" Valor {data[columna].value} columna ({columna}) ")
                    
            setattr(model, propietat, value)            

        model.save()

    carregats += [f"{file_name}"]

    if settings.DEBUG:
        amb_errors = " amb errors" if bool(errors) else "sense errors"
        print(            
            f"**carregat {file_name} ({amb_errors})**")

    return carregats, errors

def mostra_resultats():
    
    if settings.DEBUG:
        
        dades_provincia = list(
            Provincia
            .objects            
            .values()
        )        
        print(f"""
        ****************************************************************
        DADES Provincia.
        ****************************************************************
        """)
        print(json.dumps(dades_provincia, indent=4, sort_keys=True))
        
        dades_comarca = list(
            Comarca
            .objects            
            .values()
        )        
        print(f"""
        ****************************************************************
        DADES Comarca.
        ****************************************************************
        """)
        print(json.dumps(dades_comarca, indent=4, sort_keys=True))

        dades_municipi = list(
            Municipi
            .objects            
            .values()
        )        
        print(f"""
        ****************************************************************
        DADES Municipi.
        ****************************************************************
        """)
        print(json.dumps(dades_municipi, indent=4, sort_keys=True))        
        